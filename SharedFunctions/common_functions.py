import os
from openpyxl import load_workbook
import xlsxwriter
import datetime
import pandas as pd


##################################################################################################################
# Function Name: search_file
# Description  : Traverses the directory and search the file
# @param       : Directory, which you would like to traverse
# @param       : File, which you would like to search
# @return      : List of the File Location(s)
##################################################################################################################
def search_file(directory, file_name_or_extension):
    # Initialise List
    file_location_list = []

    extension = file_name_or_extension.lower()

    for dir_path, dir_names, files in os.walk(directory):
        for name in files:
            if extension and name.endswith(file_name_or_extension):
                # print(os.path.join(dir_path, name))
                file_location_list.append(os.path.join(dir_path, name))
                return file_location_list

            elif not extension:
                # print(os.path.join(dir_path, name))
                file_location_list.append(os.path.join(dir_path, name))
                return file_location_list


##################################################################################################################


##################################################################################################################
# Function Name: strip_list
# Description  : Take a list of string objects and return the same list stripped of extra whitespace.
# @param       : List which needs to be stripped of the extra white spaces
# @return      : The same list, which was given as input but stripped of whitespaces
##################################################################################################################
def strip_list(list_input):
    return [x.strip() for x in list_input]


##################################################################################################################


##################################################################################################################
# Function Name: drop_even_or_odd_elements_of_list
# Description  : Deletes the nth element of the list
# @param       : List
# @param       : Argument as Even or Odd. Zero is considered even
# @return      : The same list, with dropped indexes
##################################################################################################################
def drop_even_or_odd_elements_of_list(input_list, even_or_odd):
    if even_or_odd == "Even":
        final_list = [item for index, item in enumerate(input_list) if index % 2 == 0]
    elif even_or_odd == "Odd":
        final_list = [item for index, item in enumerate(input_list) if index % 2 != 0]
    else:
        raise Exception("Custom Error: Argument can only be Even or Odd")

    return final_list

##################################################################################################################


##################################################################################################################
# Function Name: check_folder
# Description  : Checks if the given folder exists, if not then creates a new one
# @param       : Folder Location along with Folder Name.
##################################################################################################################
def check_folder(folder_loc):
    if not os.path.exists(folder_loc):
        os.makedirs(folder_loc)

##################################################################################################################


##################################################################################################################
# Function Name: create_curr_date_folder_name
# Description  : Creates a Folder Name of Current Date in format YYYYMMDD
# @return      : String of CurrentDate Folder Name
##################################################################################################################
def create_curr_date_folder_name():
    now = datetime.datetime.now()
    year = str(now.year)
    month = now.strftime('%m')
    day = now.strftime('%d')
    str_curr_folder = year + month + day

    return str_curr_folder

##################################################################################################################


##################################################################################################################
# Function Name: write_to_excel
# Description  : Creates/Modifies the given excel
# @param       : String: sheet_name
# @param       : Dataframe: output_excel_df - dataframe, which will be written to excel
# @param       : String: output_excel_loc - Location of the excel file
# @param       : Boolean :append - True, to overwrite the existing excel file
# @return      : Null
##################################################################################################################
def write_to_excel(sheet_name, output_excel_df, output_excel_loc, append=True):
    # Check if Excel File exists
    if not os.path.isfile(output_excel_loc):
        workbook = xlsxwriter.Workbook(output_excel_loc)
        workbook.close()

    # Add data to current excel
    if append:
        workbook = load_workbook(output_excel_loc)
        # Check if Sheet Exists
        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])

        # Remove "Sheet1" from Excel
        if "Sheet1" in workbook.sheetnames:
            workbook.remove(workbook["Sheet1"])

        # Now write the workbook to excel
        writer = pd.ExcelWriter(output_excel_loc, engine='openpyxl')
        writer.book = workbook

    # Overwrite current excel
    else:
        writer = pd.ExcelWriter(output_excel_loc)

    # Write to Excel
    output_excel_df.to_excel(writer, sheet_name, index=False)
    writer.save()
    writer.close()

##################################################################################################################
