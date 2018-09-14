import os
from openpyxl import load_workbook
import xlsxwriter

##################################################################################################################
# Function Name: searchFile
# Description  : Traverses the directory and search the file 
# @param       : Directory, which you would like to traverse
# @param       : File, which you would like to search
# @return      : List of the File Location(s)
##################################################################################################################

def searchFile(Directory, FileNameOrExtenstion):
    #Initilise List
    fileLocationList = []

    extension = FileNameOrExtenstion.lower()

    for dirpath, dirnames, files in os.walk(Directory):
        for name in files:
            if extension and name.endswith(FileNameOrExtenstion):
                # print(os.path.join(dirpath, name))
                fileLocationList.append(os.path.join(dirpath, name))
                return fileLocationList

            elif not extension:
                # print(os.path.join(dirpath, name))
                fileLocationList.append(os.path.join(dirpath, name))
                return fileLocationList
            
##################################################################################################################



##################################################################################################################
# Function Name: stripList
# Description  : Take a list of string objects and return the same list stripped of extra whitespace.
# @param       : List which needs to be stripped of the extra white spaces
# @return      : The same list, which was given as input but stripped of whitespaces
##################################################################################################################
def stripList(listInput):
    return([x.strip() for x in listInput])

##################################################################################################################


##################################################################################################################
# Function Name: dropEvenOrOddElmentsOfList
# Description  : Deletes the nth element of the list
# @param       : List
# @param       : Argument as Even or Odd. Zero is considered even
# @return      : The same list, with dropped indexes
##################################################################################################################
def dropEvenOrOddElmentsOfList(list, EvenOrOdd):
    finalList = []
    if(EvenOrOdd == "Even"):
        finalList = [item for index, item in enumerate(list) if index % 2 == 0]
    elif (EvenOrOdd == "Odd"):
        finalList = [item for index, item in enumerate(list) if index % 2 != 0]
    else:
        raise Exception("Custom Error: Argument can only be Even or Odd")

    return finalList

##################################################################################################################



##################################################################################################################
# Function Name: checkFolder
# Description  : Checks if the given folder exists, if not then creates a new one
# @param       : Folder Location along with Folder Name.
##################################################################################################################
def checkFolder (folderLoc):
    if not os.path.exists(folderLoc):
        os.makedirs(folderLoc)
        
##################################################################################################################



##################################################################################################################
# Function Name: createCurrDateFolderName
# Description  : Creates a Folder Name of Current Date in format YYYYMMDD
# @return      : String of CurrentDate Folder Name
##################################################################################################################
def createCurrDateFolderName():
    now = datetime.datetime.now()
    year = str(now.year)
    month = now.strftime('%m')
    day = now.strftime('%d')
    strCurrFolder = year+month+day
    return strCurrFolder
        
##################################################################################################################


##################################################################################################################
# Function Name: write_to_excel
# Description  : Creates/Modifies the given excel
# @param       : String: sheet_name
# @param       : Dataframe: output_excel_df - dataframe, which will be written to excel
# @param       : String: output_excel_loc - Location of the excel file
# @param       : Boolean :append - True, to overwrite the existing excel file
# @return      : Null
##################################################################################################################
def write_to_excel(sheet_name, output_excel_df, output_excel_loc, append=True):
    # Check if Excel File exists
    if not os.path.isfile(output_excel_loc):
        workbook = xlsxwriter.Workbook(output_excel_loc)
        workbook.close()

    # Add data to current excel
    if append:
        workbook = load_workbook(output_excel_loc)
        # Check if Sheet Exists
        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])

        # Remove "Sheet1" from Excel
        if "Sheet1" in workbook.sheetnames:
            workbook.remove(workbook["Sheet1"])

        # Now write the workbook to excel
        writer = pd.ExcelWriter(output_excel_loc, engine='openpyxl')
        writer.book = workbook

    # Overwrite current excel
    else:
        writer = pd.ExcelWriter(output_excel_loc)

    # Write to Excel
    output_excel_df.to_excel(writer, sheet_name, index=False)
    writer.save()
    writer.close()

##################################################################################################################
