import os
from openpyxl import load_workbook
import xlsxwriter
import datetime
import pandas as pd
import win32com.client as win32     # pip install pywin32


##################################################################################################################
# Function Name: search_file
# Description  : Traverses the directory and search the file
# @param       : Directory, which you would like to traverse
# @param       : File, which you would like to search
# @return      : List of the File Location(s)
##################################################################################################################
def search_file(directory: str, file_name_or_extension: str) -> list:
    # Initialise List
    file_location_list = []

    extension = file_name_or_extension.lower()

    for dir_path, dir_names, files in os.walk(directory):
        for name in files:
            if extension and name.endswith(file_name_or_extension):
                # print(os.path.join(dir_path, name))
                file_location_list.append(os.path.join(dir_path, name))
                return file_location_list

            elif not extension:
                # print(os.path.join(dir_path, name))
                file_location_list.append(os.path.join(dir_path, name))
                return file_location_list


##################################################################################################################


##################################################################################################################
# Function Name: strip_list
# Description  : Take a list of string objects and return the same list stripped of extra whitespace.
# @param       : List which needs to be stripped of the extra white spaces
# @return      : The same list, which was given as input but stripped of whitespaces
##################################################################################################################
def strip_list(list_input: list) -> list:
    return [x.strip() for x in list_input]


##################################################################################################################


##################################################################################################################
# Function Name: drop_even_or_odd_elements_of_list
# Description  : Deletes the nth element of the list
# @param       : List
# @param       : Boolean Argument as Even = True;  Odd is False. Zero is considered even
# @return      : The same list, with dropped indexes
##################################################################################################################
def drop_even_or_odd_elements_of_list(input_list: list, even: bool) -> list:
    if even:
        final_list = [item for index, item in enumerate(input_list) if index % 2 == 0]
    else:
        final_list = [item for index, item in enumerate(input_list) if index % 2 != 0]

    return final_list

##################################################################################################################


##################################################################################################################
# Function Name: check_folder
# Description  : Checks if the given folder exists, if not then creates a new one
# @param       : Folder Location along with Folder Name.
##################################################################################################################
def check_folder(folder_loc:str):
    if not os.path.exists(folder_loc):
        os.makedirs(folder_loc)

##################################################################################################################


##################################################################################################################
# Function Name: create_curr_date_folder_name
# Description  : Creates a Folder Name of Current Date in format YYYYMMDD
# @return      : String of CurrentDate Folder Name
##################################################################################################################
def generate_curr_date_folder_name() -> str:
    now = datetime.datetime.now()
    year = str(now.year)
    month = now.strftime('%m')
    day = now.strftime('%d')
    str_curr_folder = year + month + day

    return str_curr_folder

##################################################################################################################


##################################################################################################################
# Function Name: write_to_excel
# Description  : Creates/Modifies the given excel
# @param       : String: sheet_name
# @param       : Dataframe: output_excel_df - dataframe, which will be written to excel
# @param       : String: output_excel_loc - Location of the excel file
# @param       : Boolean :append - True, to overwrite the existing excel file
# @return      : Null
##################################################################################################################
def write_to_excel(sheet_name: str, output_excel_df: pd.DataFrame, output_excel_loc: str, append: bool=True):
    # Check if Excel File exists
    if not os.path.isfile(output_excel_loc):
        workbook = xlsxwriter.Workbook(output_excel_loc)
        workbook.close()

    # Add data to current excel
    if append:
        workbook = load_workbook(output_excel_loc)
        # Check if Sheet Exists
        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])

        # Remove "Sheet1" from Excel
        if "Sheet1" in workbook.sheetnames:
            workbook.remove(workbook["Sheet1"])

        # Now write the workbook to excel
        writer = pd.ExcelWriter(output_excel_loc, engine='openpyxl')
        writer.book = workbook

    # Overwrite current excel
    else:
        writer = pd.ExcelWriter(output_excel_loc)

    # Write to Excel
    output_excel_df.to_excel(writer, sheet_name, index=False)
    writer.save()
    writer.close()

##################################################################################################################


##################################################################################################################
# Function Name: set_active_sheet
# Description  : It makes the given sheet as active
# @param       : String: sheet_name
# @param       : String: output_excel_loc - Location of the excel file
# @return      : Null
##################################################################################################################
def set_active_sheet(sheet_name: str, output_excel_loc: str):
    # Load Excel Sheet
    xls_book = load_workbook(output_excel_loc)

    # Find the index of the given sheet
    for sheet_index in range(len(xls_book.sheetnames)):
        if xls_book.sheetnames[sheet_index] == sheet_name:
            break

    # Set given sheet as active
    xls_book.active = sheet_index

    # Save the workbook
    xls_book.save(output_excel_loc)

##################################################################################################################


##################################################################################################################
# Function Name: auto_fit_columns_of_excel
# Description  : Auto Fits the column of Excel
# @param       : String: output_excel_loc - Location of excel which needs to be formatted
# @return      : Null
##################################################################################################################
def auto_fit_columns_of_excel(output_excel_loc: str):
    # Initialize
    excel = win32.gencache.EnsureDispatch('Excel.Application')

    # Read Excel
    wb = excel.Workbooks.Open(output_excel_loc)

    # Loop through all the sheets
    for sheet in wb.Worksheets:
        ws = wb.Worksheets(sheet.Name)
        ws.Columns.AutoFit()

    # Save the Excel
    wb.Save()
    wb.Close(True)
    excel.Application.Quit()

##################################################################################################################
